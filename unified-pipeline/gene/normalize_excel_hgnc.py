"""
Normalize `New_Gene_updated.xlsx` to HGNC-canonical nomenclature.

Two outputs:

1. `New_Gene_updated.xlsx.hgnc_normalized` — deterministic renames applied.
   The `Gene` column becomes the approved HGNC symbol; the original
   clinical name moves to `Gene Aliases` if not already there. Safe
   cases only — edits nothing ambiguous.

2. `dictionaries/pending_excel_review.csv` — rows that the script
   *cannot* safely normalize. Tom's one-time manual cleanup list:
   gene families (NTRK 1/2/3), ambiguous markers (p40, TTF-1),
   mutation-specific rows (TP53 p.R248Q), chromosomal segments
   (1p19q), etc.

Run:
    python normalize_excel_hgnc.py                 # default paths
    python normalize_excel_hgnc.py --in X --out Y  # override
"""

import argparse
import csv
import os
import shutil
import sys
from typing import Dict, List, Optional, Tuple

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)

try:
    import openpyxl
except ImportError as e:
    raise SystemExit(
        "openpyxl required. Install: pip install openpyxl"
    ) from e

DEFAULT_IN = os.path.join(_HERE, "New_Gene_updated.xlsx")
DEFAULT_OUT = os.path.join(_HERE, "New_Gene_updated.hgnc_normalized.xlsx")
PENDING_CSV = os.path.join(_HERE, "dictionaries", "pending_excel_review.csv")


# --- Deterministic renames. All verified against current HGNC registry. ---
#
# Tuples: (old_gene_name, new_hgnc_symbol, alias_to_preserve_or_None)
# alias_to_preserve=None means the rename drops the old name entirely (e.g.
# "NTRK 1" → "NTRK1"; "NTRK 1" isn't a useful alias to keep).
SAFE_RENAMES: List[Tuple[str, str, Optional[str]]] = [
    # Clinical name → HGNC, keep clinical name as alias
    ("PD-1",            "PDCD1",  "PD-1"),
    ("HER2 (ERBB2)",    "ERBB2",  "HER2"),
    ("STK11 (LKB1)",    "STK11",  "LKB1"),
    ("FANCN (PALB2)",   "PALB2",  "FANCN"),   # FANCN is deprecated

    # Spacing / formatting cleanup — no alias worth preserving
    ("NTRK 1",          "NTRK1",  None),
    ("NTRK 2",          "NTRK2",  None),
    ("NTRK 3",          "NTRK3",  None),

    # Unicode / case
    ("PPARγ",           "PPARG",  "PPARγ"),
]

# --- Rows that need Tom's manual review (not auto-renamed). ---
PENDING_REVIEW: List[Tuple[str, str]] = [
    # (row_value, reason)
    ("1p19q",           "chromosomal co-deletion, not a single HGNC gene"),
    ("H3-3A",           "valid HGNC — leave as-is? (double-check)"),
    ("LANA-1",          "KSHV viral antigen, not a human HGNC gene"),
    ("TDt",             "probably DNTT (terminal deoxynucleotidyl transferase)"),
    ("IGHV4-34",        "valid HGNC immunoglobulin variable segment — leave as-is?"),
    ("IgH",             "should be IGH (immunoglobulin heavy locus)?"),
    ("CDKN2A/B",        "gene pair — split into CDKN2A + CDKN2B, or keep grouped?"),
    ("HERV-K",          "endogenous retrovirus, not HGNC"),
    ("Napsin A",        "IHC marker (NAPSA) — split or keep clinical name?"),
    ("TTF-1",           "IHC marker (NKX2-1) — split or keep?"),
    ("p40",             "IHC marker (TP63 isoform) — collapse to TP63?"),
    ("p63",             "IHC marker (TP63) — rename to TP63?"),
    ("NTRK 1/2/3",      "gene family — split into NTRK1/2/3 or remove (already exist)?"),
    ("MTCP1-B1",        "check — MTCP1 is HGNC but MTCP1-B1 isn't standard"),
    ("IKZF1 p.N159Y",   "mutation-specific row — separate entry by design?"),
    ("IKZF1plus",       "risk classification label, not a gene — keep?"),
    ("TP53 p.R248Q",    "mutation-specific row — separate entry by design?"),
    ("ZEB2 p.H1038R",   "mutation-specific row — separate entry by design?"),
    ("BLM/RECQL3",      "BLM is HGNC; RECQL3 is legacy name for BLM — merge?"),
    ("IDH1/2",          "gene pair — split into IDH1 + IDH2?"),
]


def normalize(src: str, dst: str) -> Dict:
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # Find header row columns
    rows_iter = ws.iter_rows(values_only=False)
    header_cells = next(rows_iter)
    header = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    # Expected columns: Gene, Gene Aliases, Broad Cancers
    try:
        gene_col = header.index("Gene")
    except ValueError:
        raise SystemExit(f"No 'Gene' column in header: {header}")
    try:
        alias_col = header.index("Gene Aliases")
    except ValueError:
        alias_col = None

    rename_map = {old: (new, alias) for (old, new, alias) in SAFE_RENAMES}
    applied: List[Tuple[str, str]] = []

    # Walk data rows and apply safe renames in place.
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        gene_cell = row[gene_col]
        gene_val = (gene_cell.value or "")
        gene_val_str = str(gene_val).strip()
        if gene_val_str not in rename_map:
            continue
        new_gene, preserved_alias = rename_map[gene_val_str]
        gene_cell.value = new_gene

        # Merge the old clinical name into aliases (if requested).
        if preserved_alias and alias_col is not None:
            alias_cell = row[alias_col]
            existing = str(alias_cell.value or "").strip()
            alias_list = [a.strip() for a in existing.split(",") if a.strip()]
            if preserved_alias not in alias_list:
                alias_list.insert(0, preserved_alias)
            alias_cell.value = ", ".join(alias_list)

        applied.append((gene_val_str, new_gene))

    wb.save(dst)
    return {"applied": applied, "total_renames": len(applied)}


def write_pending(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["excel_gene_value", "reason_or_suggestion",
                    "decision", "tom_notes"])
        for old, reason in PENDING_REVIEW:
            w.writerow([old, reason, "", ""])


def main(src: str, dst: str) -> None:
    if not os.path.exists(src):
        raise SystemExit(f"input not found: {src}")
    print(f"[normalize] src={src}")
    print(f"[normalize] dst={dst}")
    stats = normalize(src, dst)
    print(f"[normalize] applied {stats['total_renames']} safe renames:")
    for old, new in stats["applied"]:
        print(f"    {old!r:30} → {new}")

    write_pending(PENDING_CSV)
    print(f"[normalize] wrote pending-review list → {PENDING_CSV}")
    print(f"  {len(PENDING_REVIEW)} ambiguous rows for Tom to decide")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", default=DEFAULT_IN)
    ap.add_argument("--out", dest="dst", default=DEFAULT_OUT)
    args = ap.parse_args()
    main(args.src, args.dst)
